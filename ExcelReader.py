from openpyxl import load_workbook as work_book


class ExcelReader:
    def __init__(self, file_name):
        self.fileName = file_name
        self.table = []
        wb = work_book(self.fileName, True)
        head = True
        for sheet in wb.worksheets:
            for row in sheet.rows:
                if head:
                    head = False
                    continue
                values = []
                for cell in row:
                    if cell.value is not None:
                        values.append(cell.value)
                    else:
                        values.append("")
                self.table.append(values)

    @staticmethod
    def save_to_file(file_name, this_address):
        wb = work_book(file_name, False)
        sheet = wb.worksheets[0]
        id_row = 0
        for row in this_address:
            id_row += 1
            sheet.cell(None, id_row, 1, str(row.region))
            sheet.cell(None, id_row, 2, str(row.city))
            sheet.cell(None, id_row, 3, str(row.address))
            sheet.cell(None, id_row, 4, str(row.index))
            sheet.cell(None, id_row, 5, str(row.find_address))
            sheet.cell(None, id_row, 6, str(row.centroid))

        wb.save(file_name)

    def __del__(self):
        self.table = None
